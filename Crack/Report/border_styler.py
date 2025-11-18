from openpyxl.styles import Border, Side

class BorderStyler:
    def __init__(self):
        # 테두리 스타일 정의
        self.thin = Side(border_style="thin", color="000000")  # 얇은 테두리
        self.thick = Side(border_style="thick", color="000000")  # 두꺼운 테두리
        self.THIN = Border(
            top=self.thin, bottom=self.thin, left=self.thin, right=self.thin
        )
        self.THICK = Border(
            top=self.thick, bottom=self.thick, left=self.thick, right=self.thick
        )

    # [표지] 굵은 테두리
    def cover_thick_border(self, sheet, rowTop, rowBot, colLeft, colRight):
        rows = range(rowTop, rowBot + 1)
        cols = range(colLeft, colRight + 1)

        for row in rows:
            for col in cols:
                cell = sheet.cell(row=row, column=col)
                current_border = cell.border  # 기존 테두리 가져오기

                # 외부 테두리만 두껍게 설정
                new_border = Border(
                    left=self.THICK.left if col == colLeft else current_border.left,
                    right=self.THICK.right if col == colRight else current_border.right,
                    top=self.THICK.top if row == rowTop else current_border.top,
                    bottom=self.THICK.bottom if row == rowBot else current_border.bottom
                )
                cell.border = new_border
                
    # [표지] 얇은 테두리
    def cover_thin_border(self, sheet, cell_range):
        # 지정된 범위의 셀에 얇은 테두리 적용
        for row in sheet[cell_range]:
            for cell in row:
                cell.border = self.THIN

    # [속지] 굵은 테두리
    def part_thick_border(self, sheet, rowTop, rowBot, colLeft, colRight):
        rows = range(rowTop, rowBot + 1)
        cols = range(colLeft, colRight + 1)
        
        for row in rows:
            for col in cols:
                cell = sheet.cell(row=row, column=col)
                current_border = cell.border  # 기존 테두리 가져오기

                # 외부 테두리만 두껍게 설정
                new_border = Border(
                    left=self.THICK.left if col == colLeft else current_border.left,
                    right=self.THICK.right if col == colRight else current_border.right,
                    top=self.THICK.top if row == rowTop else current_border.top,
                    bottom=self.THICK.bottom if row == rowBot else current_border.bottom
                )
                cell.border = new_border
    # [속지] 얇은 테두리
    def part_inner_border(self, sheet, rowTop, rowBot, colLeft, colRight):
        rows = range(rowTop, rowBot + 1)
        cols = range(colLeft, colRight + 1)

        for row in rows:
            for col in cols:
                cell = sheet.cell(row=row, column=col)
                current_border = cell.border  # 기존 테두리 가져오기

                # 외부 테두리는 굵게, 내부 테두리는 얇게 설정
                new_border = Border(
                    left=self.THICK.left if col == colLeft else self.THIN.left,
                    right=self.THICK.right if col == colRight else self.THIN.right,
                    top=self.THICK.top if row == rowTop else self.THIN.top,
                    bottom=self.THICK.bottom if row == rowBot else self.THIN.bottom
                )
                cell.border = new_border
