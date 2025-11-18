import time
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Color
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

import json
import cv2
import math
import os
import numpy as np
import datetime
import os

from Crack.Report.border_styler import BorderStyler
from Crack.Report.crack_heatmap import CrackMap
from Crack.Report.image_split import ImageSplit

class report: 
    def __init__(self, bridgeInfo, imagePath, jsonPath, save_dir):
        self._bridgeInfo = bridgeInfo
        self._imagePath = imagePath
        self._jsonPath = jsonPath
        self._saveDir = save_dir

    # <----------------------------------- 함수 실행 ------------------------------------>
    def run(self):
        split = ImageSplit(self._imagePath, self._jsonPath) 
        list = split.SplitImage()
        self.generateExcel(list, self._imagePath, self._jsonPath, self._saveDir)

    # <------------------------------- 보고서 인쇄영역 설정 -------------------------------->
    def set_print_settings(self, sheet):
        sheet.print_area = 'A1:U39'
        
        # 용지 방향 및 크기 설정
        sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A3
        
        # fitToWidth, fitToHeight을 조정하여 페이지 크기 맞춤 설정
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 1  
        
        # 여백을 좁은 여백으로 설정
        sheet.page_margins.left = 0.25
        sheet.page_margins.right = 0.25
        sheet.page_margins.top = 0.25
        sheet.page_margins.bottom = 0.25
        sheet.page_margins.header = 0.25
        sheet.page_margins.footer = 0.25
        
        # 인쇄 배율을 100%로 설정하여 크기 조절을 명확히 함
        sheet.page_setup.scale = 100

        # 상하좌우 가운데 정렬 설정
        sheet.print_options.horizontalCentered = True
        sheet.print_options.verticalCentered = True

    # <---------------------------------- 보고서 생성 ----------------------------------->        
    def generateExcel(self, data_list, image_path, json_data_path, save_dir): 
        bridgeInfo = self._bridgeInfo
        wb = Workbook()  # Workbook 생성
        wb.remove(wb['Sheet'])  # 빈 시트 삭제
        
        # BorderStyler 인스턴스 설정
        border_styler = BorderStyler()
        
        # 액셀 스타일 지정 (폰트크기, 셀색상, 가운데정렬등)
        style_center = Alignment(horizontal='center', vertical='center')
        cover_title_font = Font(size=24, bold=True, color='000000')
        title_font = Font(size=18, bold=True, color='000000')
        title_content_font = Font(size=18, bold=False, color='000000')
        subject_font = Font(size=12, bold=True, color="000000")
        fill_color = PatternFill(fill_type='solid', fgColor=Color('e9e9e9'))
        
        # 셀 너비 보정함수
        def set_column_width(ws, col, width):
            ws.column_dimensions[col].width = width * 1.2  
                
        # 셀 높이 보정함수
        def set_row_height(ws, row, height):
            ws.row_dimensions[row].height = height * 1.2  
            
        # 커버에 들어가는 이미지 분할 함수 (6장)
        def split_image(image_path, split_count=6):
            # img = cv2.imread(image_path)
            img_array = np.fromfile(image_path, np.uint8)
            img = cv2.imdecode(img_array, cv2.IMREAD_COLOR)

            if img is None:
                print("❌ [커버이미지 6장 분할] : 이미지를 불러올 수 없습니다.")
                return []
            
            height, width, _ = img.shape
            split_height = math.ceil(height / split_count)
            
            # 저장 폴더 설정
            dir_path = os.path.dirname(image_path)
            file_name = os.path.basename(image_path).split('.')[0].replace("_result", "")
            save_dir = os.path.join(dir_path, f"{file_name}_split")
            os.makedirs(save_dir, exist_ok=True)
            
            split_paths = []
            for i in range(split_count):
                start_y = i * split_height
                end_y = min((i + 1) * split_height, height)
                split_img = img[start_y:end_y, :]
                
                save_path = os.path.join(save_dir, f"{file_name}_cover_split_{i+1}.png")

                ext = os.path.splitext(save_path)[1]
                result, n = cv2.imencode(ext, split_img, None)
                if result:
                    with open(save_path, mode='w+b') as f:
                        n.tofile(f)

                split_paths.append(save_path)
            
            return split_paths    
        
        # 이미지 사이즈를 비율에 따라 변경하고 셀 내에 배치하는 함수 
        def image_ratio_resize_arrangement (sheet, sheet_img, target_width, target_height, target_y, target_x) :
            original_width, original_height = sheet_img.width, sheet_img.height
            
            # 비율 유지하며 크기 조정
            if original_width > target_width:
                sheet_img.width = target_width
                sheet_img.height = int(original_height * (target_width / original_width))  # 비율 유지
            if sheet_img.height > target_height:
                sheet_img.height = target_height
                sheet_img.width = int(original_width * (target_height / original_height))  # 비율
                
            col_offset = target_y - round(sheet_img.width / 2 / 68)
            row_offset = target_x - round(sheet_img.height / 2 / 23)
            col_offset = max(1, col_offset)
            row_offset = max(1, row_offset)    
            final_col = get_column_letter(int(col_offset))
        
            sheet_img.anchor = f"{final_col}{row_offset}"
            sheet.add_image(sheet_img)
            
        # <------------------------------- 보고서 출력 [표지] --------------------------------> 
        cover_sheet = wb.create_sheet(title="cover", index=0)
        self.set_print_settings(cover_sheet)  # cover_sheet에 대한 인쇄 설정 적용
        
        # 테두리 스타일 정의
        border_styler.cover_thick_border(cover_sheet, 1, 8, 1, 16)
        border_styler.cover_thin_border(cover_sheet, 'A1:U39')
        
        # [표지 셀 너비 설정]
        for col in range(1, 7):  # A부터 P까지
            column_letter = get_column_letter(col)
            set_column_width(cover_sheet, column_letter, 9.5) 
        for col in range(7, 22):  # A부터 P까지
            column_letter = get_column_letter(col)
            set_column_width(cover_sheet, column_letter, 6.5) 
        
        # [표지] 병합
        cover_sheet.merge_cells('A1:U3')
        cover_sheet.merge_cells('A4:C5')
        cover_sheet.merge_cells('D4:F5')
        cover_sheet.merge_cells('G4:U5')
        cover_sheet.merge_cells('A6:C39')
        cover_sheet.merge_cells('D6:F39')
        
        cover_sheet.merge_cells('G6:K20')
        cover_sheet.merge_cells('G21:K22')
        cover_sheet.merge_cells('L6:P20')
        cover_sheet.merge_cells('L21:P22')
        cover_sheet.merge_cells('Q6:U20')
        cover_sheet.merge_cells('Q21:U22')
        
        cover_sheet.merge_cells('G23:K37')
        cover_sheet.merge_cells('G38:K39')
        cover_sheet.merge_cells('L23:P37')
        cover_sheet.merge_cells('L38:P39')
        cover_sheet.merge_cells('Q23:U37')
        cover_sheet.merge_cells('Q38:U39')
        
        # [표지] 텍스트 적용되는 셀
        cover_sheet_title_range = ['A1', 'A4', 'D4', 'G4', 'G21', 'L21', 'Q21', 'G38', 'L38', 'Q38']
        
        # [표지] 전체 셀 가운데 정렬
        for row in cover_sheet.iter_rows(min_row=1, max_row=40, min_col=1, max_col=21):
            for cell in row:
                cell.alignment = style_center
                
        # [표지] 제목 칼럼 폰트 
        for pos in cover_sheet_title_range : 
            cover_sheet[pos].font = cover_title_font
            
        # [표지]
        cover_sheet['A1'].value = bridgeInfo['bridgeName']+('-'+bridgeInfo['pierName']+'-'+bridgeInfo['pierSideNo']+'-'+bridgeInfo['cameraNo']+'CAM').replace(' ', '')
        cover_sheet['A4'].value ="Stitching Image"
        cover_sheet['D4'].value ="Crack Heatmap"
        cover_sheet['G4'].value = "Part of " + (bridgeInfo['pierName']+'-'+bridgeInfo['pierSideNo']+'-'+bridgeInfo['cameraNo']+'CAM').replace(' ', '')
        cover_sheet['G21'].value = "Part 1 ~ 4"
        cover_sheet['G38'].value = "Part 5 ~ 8"
        cover_sheet['L21'].value = "Part 9 ~ 12"
        cover_sheet['L38'].value = "Part 13 ~ 16"
        cover_sheet['Q21'].value = "Part 17 ~ 20"
        cover_sheet['Q38'].value = "Part 21 ~ 24"
        
        # [표지] 색상이 적용되는 셀 목록
        cover_sheet_cell_to_fill = ['A1', 'A4', 'D4', 'G4', 'G21', 'L21', 'Q21', 'G38', 'L38', 'Q38']
        
        for cell in cover_sheet_cell_to_fill:
            cover_sheet[cell].fill = fill_color
                
        # 표지 전체 이미지 삽입 (1장)      
        cover_sheet_img = Image(image_path)
        image_ratio_resize_arrangement(cover_sheet, cover_sheet_img, 300, 800, 2, 25) 
        
        # 표지 히트맵 이미지 생성 및 삽입 (1장)
        crack_map = CrackMap(json_data_path, image_path)
        crack_map_img_path = crack_map.draw_crack_map(image_path)  
        cover_sheet_crackmap_img = Image(crack_map_img_path) 
        image_ratio_resize_arrangement(cover_sheet, cover_sheet_crackmap_img, 300, 800, 6, 25)

        # 표지 분할 이미지 삽입 (6장)
        splited_cover_image = split_image(image_path)

        # 분할 이미지(6장) 삽입 위치를 리스트로 정의, 위치는 가운데 정렬하려고 대충 눈대중으로 맞춘거임
        positions = [(9, 14), (14, 14), (19, 14), (9, 31), (14, 31), (19, 31)]
        for index, data in enumerate(splited_cover_image):
            if index < len(positions):  # positions 리스트의 길이 체크
                v = Image(data)
                image_ratio_resize_arrangement(cover_sheet, v, 200, 300, *positions[index])

        # <----------------------------------- 보고서 출력 [속지] ------------------------------------> 
        for index, data in enumerate(data_list):
            part_sheet = wb.create_sheet(title=data["part"])
            self.set_print_settings(part_sheet)  # part_sheet에 대한 인쇄 설정 적용
            
            # 테두리 스타일 정의
            border_styler.cover_thin_border(part_sheet, 'A1:U39')
            border_styler.part_thick_border(part_sheet, 1, 8, 1, 16)
            border_styler.part_thick_border(part_sheet, 9, 39, 1, 16)
            border_styler.part_thick_border(part_sheet, 1, 39, 18, 21)

            # 셀 너비 설정 - A~P 열 너비 설정 (보정 적용)
            for col in range(1, 17):  # A부터 P까지
                column_letter = get_column_letter(col)
                set_column_width(part_sheet, column_letter, 6)  

            # 셀 높이 설정 1~39 행 높이 설정 (보정 적용)
            for row in range(1, 40):  
                set_row_height(part_sheet, row, 16.5)  
                set_row_height(cover_sheet, row, 16.8) 

            # 개별 열 너비 설정 (보정 적용)
            set_column_width(part_sheet, 'Q', 2)
            set_column_width(part_sheet, 'R', 8)
            set_column_width(part_sheet, 'S', 14)
            set_column_width(part_sheet, 'T', 22)
            set_column_width(part_sheet, 'U', 22)

            # 병합
            part_sheet.merge_cells('A1:C2')
            part_sheet.merge_cells('A3:C4')
            part_sheet.merge_cells('A5:C6')
            part_sheet.merge_cells('A7:C8')
            part_sheet.merge_cells('D1:H2')
            part_sheet.merge_cells('D3:H4')
            part_sheet.merge_cells('D5:H6')
            part_sheet.merge_cells('D7:P8')
            part_sheet.merge_cells('I1:K2')
            part_sheet.merge_cells('I3:K4')
            part_sheet.merge_cells('I5:K6')
            part_sheet.merge_cells('L1:P2')
            part_sheet.merge_cells('L3:P4')
            part_sheet.merge_cells('L5:P6')
            part_sheet.merge_cells('A9:P39')
            part_sheet.merge_cells('Q1:Q39')

            # 텍스트 적용되는 셀
            part_sheet_title_range = ['A1','I1','A3','I3','A5','I5','A7']
            part_sheet_title_content_range = ['D1','L1','D3','L3','D5','L5','D7']
            part_sheet_subject_range = ['R1','S1','T1','U1']

            local_time = time.localtime()
            formatted_time = time.strftime('%Y-%m-%d', local_time)

            # now = datetime.datetime.now()
            # current_date_str = now.strftime('%Y-%m-%d')

            # 전체 셀 가운데 정렬
            for row in part_sheet.iter_rows(min_row=1, max_row=40, min_col=1, max_col=21):
                for cell in row:
                    cell.alignment = style_center
            
            # 제목 칼럼 폰트 
            for pos in part_sheet_title_range : 
                part_sheet[pos].font = title_font
                
            # 제목 컨텐츠 폰트
            for pos in part_sheet_title_content_range :
                part_sheet[pos].font = title_content_font
            
            # 우측 테이블 칼럼 폰트
            for pos in part_sheet_subject_range : 
                part_sheet[pos].font = subject_font
                
            part_sheet['A1'].value = 'BRIDGE'
            part_sheet['D1'].value = bridgeInfo['bridgeName']
            part_sheet['I1'].value = 'DATE'
            part_sheet['L1'].value = formatted_time
            part_sheet['A3'].value = 'DIVISION'
            part_sheet['D3'].value = bridgeInfo['pierName']
            part_sheet['I3'].value = 'SIDE'
            part_sheet['L3'].value = bridgeInfo['pierSideNo']
            part_sheet['A5'].value = 'CAM. NO'
            part_sheet['D5'].value = bridgeInfo['cameraNo']
            part_sheet['I5'].value = 'PART'
            part_sheet['L5'].value = data["part"].replace("part", "")
            part_sheet['A7'].value = 'DETAIL_ID'
            part_sheet['D7'].value = bridgeInfo['bridgeName']+('-'+bridgeInfo['pierName']+'-'+bridgeInfo['pierSideNo']+'-'+bridgeInfo['cameraNo']+'CAM'+'_'+data["part"]).replace(' ', '')
            part_sheet['R1'].value = 'No.'
            part_sheet['S1'].value = 'Crack_ID'
            part_sheet['T1'].value = 'Estimate_Length(mm)'
            part_sheet['U1'].value = 'Estimate_Width(mm)'

            # 색상이 적용되는 셀 목록
            part_sheet_cell_to_fill = ['A1', 'I1', 'A3', 'I3', 'A5', 'I5', 'A7', 'R1', 'S1', 'T1', 'U1']

            # 각 셀에 대해 fill 적용
            for cell in part_sheet_cell_to_fill:
                part_sheet[cell].fill = fill_color
                
            # 이미지, 테이블 데이터 삽입    
            row = 2 # R2부터 U39 셀에 배치하기 위해 행 번호 초기화

            # 각 파트를 처리
            part = data  # data_list에서 각 data 항목을 처리하는 방식으로 수정
            imagePath = part['imagePath']
            part_sheet_img = Image(imagePath)            
            image_ratio_resize_arrangement(part_sheet, part_sheet_img, 915, 750, 8, 23)  
            
            # 열 데이터 입력
            for idx, json_item in enumerate(part["jsonPath"], start=1):
                # 각 행에 데이터를 채운다
                part_sheet[f'R{row}'] = json_item['NO']
                part_sheet[f'S{row}'] = json_item['CRACK_ID']
                part_sheet[f'U{row}'] = json_item['ESTIMATE_WIDTH']
                part_sheet[f'T{row}'] = json_item['ESTIMATE_HEIGTH']

                # 행 번호 증가
                row += 1

        # <--------------------------- 파일 저장 ---------------------------->
        my_save_dir = save_dir
        bridge_name = bridgeInfo['bridgeName']+('-'+bridgeInfo['pierName']+'-'+bridgeInfo['pierSideNo']+'-'+bridgeInfo['cameraNo']+'CAM').replace(" ", "")
        save_name = '_Report.xlsx'
        save_path = os.path.join(my_save_dir, bridge_name + save_name)

        wb.save(save_path)